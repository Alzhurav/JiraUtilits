import openpyxl
import datetime

from jira.exmple import JiraExample
from jira.exmple.JiraTimeTask import JiraTimeTask

wb2 = openpyxl.load_workbook('example.xlsx')
ws = wb2.active

# print(ws['C3'].value + ' ' + ws['B3'].value + ' ' + str(ws['A3'].value))

# открыли файл
wb2 = openpyxl.load_workbook('example.xlsx')
ws = wb2.active
# идем по строкам
date = ''
dateTasks = []
for row in ws.iter_rows(row_offset=1):
    time = row[0].value
    comment = row[2].value
    jira_number = row[3].value

    if time != '':
        if type(row[0].value) is datetime.datetime:
            date = datetime.datetime.strftime(row[0].value, '%d/%m/%Y')
            # print(date)
            continue

        if comment is not None and jira_number is not None:
            dateTasks.append(JiraTimeTask(date, jira_number, time, comment))
            # print(jira_number + ':  ' + str(time) + 'm. --- ' + comment)

JiraExample.save_tasks(dateTasks)
# находим дату равную сегодняшней/можно запустить с конкретной датой
# от этой строки начинаем разбирать всё до след даты или до пустой строки
# выделяем необходимые параметры
# записываем их в задачу в jira
# логируем в какую задачу что записали
# Как только получили -> прекаратили обработку


# font0 = xlwt.Font()
# font0.name = 'Times New Roman'
# font0.colour_index = 2
# font0.bold = True
#
# style0 = xlwt.XFStyle()
# style0.font = font0
#
# style1 = xlwt.XFStyle()
# style1.num_format_str = 'D-MMM-YY'
#
# wb = xlwt.Workbook()
# ws = wb.add_sheet('A Test Sheet')
#
# ws.write(0, 0, 'Test', style0)
# ws.write(1, 0, datetime.now(), style1)
# ws.write(2, 0, 'dfsdfs')
# ws.write(2, 1, 1)
# ws.write(2, 2, xlwt.Formula("A3+B3"))
#
# wb.save('example.xls')